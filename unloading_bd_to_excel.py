import openpyxl
import sqlite3
import re


wb = openpyxl.Workbook()
if 'Students' in wb.sheetnames:
    wb.remove(wb['Students'])
wb.create_sheet(title='Students', index=0)
sheet = wb['Students']
con = sqlite3.connect('main_data_base.db')
cursor = con.cursor()
rows = cursor.execute('SELECT * FROM users_id').fetchall()
for i in range(0, len(rows)):
    for j in range(0, 3):
        if j == 0:
            sheet.cell(row=i + 1, column=j + 1).value = rows[i][2]
        elif j == 1:
            sheet.cell(row=i + 1, column=j + 1).value = rows[i][1]
        else:
            if rows[i][3] == None:
                sheet.cell(row=i + 1, column=j + 1).value = 0
            else:
                sheet.cell(row=i + 1, column=j + 1).value = round(len(re.split(' ', str(rows[i][3]))) * 100 / 30, 2)
cursor.close()
con.commit()
con.close()
wb.save('Results.xlsx')
